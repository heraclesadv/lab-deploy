import os 
import shutil
import time
from Resources.samples import *
from Resources.labClass import *
import sys
from dotenv import load_dotenv
from openpyxl import load_workbook
from Resources.Users import *


load_dotenv()

#Permet d'ajouter des VM à un LAB

def main():

    lockfile()
    

    if len(sys.argv) == 2 and (sys.argv[1] == "-h" or sys.argv[1] == "help" ):
        fichier = open("Resources/help/helpAddVM.txt", 'r')
        print(fichier.read())
        fichier.close()
        sys.exit()

    #le deuxième argument doit faire référence au nom du Lab mais
    #si il n'est pas renseigné celui-ci sera demandé 
    if len(sys.argv) == 1 :
        b = input("Please enter the name of the lab to load: ")
    else :
        b = sys.argv[1]    
    try:
        print(b)
        l = loadLab(b)
    except:
        print("Failed ! Do your lab exists ? ")
        sys.exit()

    #Demande les VMs (ubuntu, win, ... ) à ajouter ainsi que leurs roles et verifie si ils existent
    if len(sys.argv) == 3 :
        #recuperer le fichier excel ajouté en argument dans le dossier Resources 
        wb = load_workbook("Resources/"+sys.argv[2])
        ws = wb.active
        ligne = 1
        test2= True
        while test2:
            c = ""
            colone = 1
            #vérifie que l'on est à la fin de la ligne ou à la fin du fichier 
            if ws.cell(row=ligne, column= colone).value != None and ws.cell(row=ligne, column= colone).value != "":
                test= True
                while test:
                    print(ws.cell(row=ligne, column= colone).value)
                    if ws.cell(row=ligne, column= colone).value != None and ws.cell(row=ligne, column= colone).value != "" :
                        c += ws.cell(row=ligne, column= colone).value +" "
                        colone +=1
                    else :
                        test= False
                c = c[:-1]
                ca = c.split(" ")
                for d in ca:
                    if not d in ROLES:
                        print("Role "+d+" does not exist, computer not added.")
                        break
                else:
                    l.addComputer(c)
                    print("Computer added.")
                ligne +=1
            else :
                test2 =False
        wb.close()
    else :
        #Demande les VMs (ubuntu, win, ... ) à ajouter ainsi que leurs roles et verifie si ils existent
        while True:
            c = input("add >> ")
            if c.replace(" ", "") == "":
                break
            ca = c.split(" ")
            for d in ca:
                if not d in ROLES:
                    print("Role "+d+" does not exist, computer not added.")
                    break
            else:
                l.addComputer(c)
                print("Computer added.")
    
    
    print("Recreating terraform files...")
    os.chdir("..")
    os.mkdir("Labs/" + l.name + '_add/')
    shutil.copyfile("terraform/header.tf", "Labs/" + l.name + "_add/header.tf")
    
    # Pour chaque ordinateur non existants on utilise le sample correspondant et on le copy dans le dossier "_add":
    listNewComp = []
    for ordi in l.computers:
        filename = 'Labs/' + l.name + '/' + ordi.name + '.tf'
        if not os.path.exists(filename) :
            importConfigFile(TYPES[ordi.getType()][0], "Labs/" + l.name + '/' + ordi.name + '.tf', 
                {"name": ordi.name, "MACAddressHostOnly":ordi.macAddressHostOnly, "MACAddressLanPortGroup":ordi.macAddressLanPortGroup})
            shutil.copyfile("Labs/" + l.name + '/' + ordi.name + '.tf', "Labs/" + l.name + '_add/' + ordi.name + '.tf')
            listNewComp.append(ordi.name)

    # On remplit le fichier de configuration:
    importConfigFile("terraform/variableSample.tf", 
        "Labs/" + l.name + "_add/variables.tf",
        {"ESXiIP":  os.getenv('ESXi'), 
        "ESXiUser":os.getenv('user'), 
        "ESXiPwd":os.getenv('password'),
        "ESXiDatastore": os.getenv('datastore'),
        "VMNetwork":os.getenv('VMNetwork'),
        "HostOnlyNetwork":l.network.name
        })
    

    shutil.copyfile("terraform/versions.tf", "Labs/" + l.name + "_add/versions.tf")
    shutil.copytree("terraform/.terraform", "Labs/" + l.name + "_add/.terraform")
    os.chdir("Script")


    if not ask("Files created, go ? It's the right moment to edit the terraform files !"):
        print("Aborting...")
        l.network.freeHostOnlyNetwork()
        shutil.rmtree("../Labs/" + l.name +"_add")
        os.system("rm ../lock")
        return None
                    
    print("Running Terraform ...")
    os.chdir("../Labs/"+l.name+'_add/')
    os.system("terraform init")
    os.system("terraform apply -auto-approve")
    os.chdir("../../Script")
    time.sleep(15)
    # Crée les fichiers de configuration ansible à partir des samples dans ansible/roles/samples et ex nihilo
    print("Creating ansible files...")
    l.cleanAnsibleFiles()
    # fichier de configuration de guacamole qui est téléversé sur le logger par ansible
    l.buildGuacamoleConfigFile()
    # On crée ex nihilo le fichier inventory.yml
    os.system("rm ../ansible/inventory.yml")
    fichier = open("../ansible/inventory.yml", 'a')
    guacaComp=l.name+"ubuntu0"
    mask = l.network.getIPmask()

    for ordi in l.computers:
        id = os.popen("sshpass -p " + os.getenv('password')+ " ssh -o StrictHostKeyChecking=no " + os.getenv('user') + "@" + os.getenv('ESXi') + " " + "vim-cmd vmsvc/getallvms | grep \"" + ordi.name + "/" + ordi.name + ".vmx\" | awk '{print $1}' | awk '{$1=$1};1'").read().replace("\n", "")
        
        if ordi.name in listNewComp :
            IP = os.popen("sshpass -p " + os.getenv('password')+ " ssh -o StrictHostKeyChecking=no " + os.getenv('user') + "@" + os.getenv('ESXi') + " " + "vim-cmd vmsvc/get.guest "+id+" | grep -m 1 '192.168.1' | sed 's/[^0-9+.]*//g'").read().replace("\n", "")
        else :
            IP = os.popen("sshpass -p " + os.getenv('password')+ " ssh -o StrictHostKeyChecking=no " + os.getenv('user') + "@" + os.getenv('ESXi') + " " + "vim-cmd vmsvc/get.guest "+id+" | grep -m 1 '"+mask+"' | sed 's/[^0-9+.]*//g'").read().replace("\n", "")
        ordi.ESXiID = id
        ordi.dhcpIP = IP 

    # on réunit les ordinateurs par type, car il existe un fichier group_vars par type
    for type in TYPES:
        ordis = l.getOrdiWithType(type)
        hosts= ""
        for ordi in ordis:
            hosts += "    "+ordi.dhcpIP+":\n"
        fichier.write(type + ":\n  hosts:\n" + hosts + '\n')
    fichier.close()

    # On crée ex nihilo le fichier detectionlab.yml
    os.system("rm ../ansible/detectionlab.yml")
    fichier = open("../ansible/detectionlab.yml", 'a')
    fichier.write("---\n")
    
    for ordi in l.computers:
        if ordi.name == guacaComp:
            fichier.write("- hosts: " + ordi.dhcpIP)
            fichier.write("\n  roles:\n")
            ordi.buildAnsibleTasks(l.getDNSIP()) # dans cette fonction sont crée les roles à partir des samples
            fichier.write("    - "+ "guacamoleActualise" + "\n")
            fichier.write("  tags: " + ordi.name + "\n\n")
        elif ordi.name in listNewComp :
            fichier.write("- hosts: " + ordi.dhcpIP)
            fichier.write("\n  roles:\n")
            ordi.buildAnsibleTasks(l.getDNSIP()) 
            for role in ordi.ansibleRoles:
                fichier.write("    - "+ role + "\n")
            fichier.write("  tags: " + ordi.name + "\n\n")
    fichier.close()
    
    
    
    
    
    print("Running ansible...")
    # Lance les scripts ansible créés ordinateur par ordinateur
    for ordi in l.computers:
        os.chdir("../ansible")
        os.system("ansible-playbook detectionlab.yml --tags \""+ordi.name+"\" --timeout 180")
        os.chdir("../Script")
    time.sleep(5)
    print("Done ! Cleaning...")
    l.cleanAnsibleFiles()
    print("Disconnecting management network...")
    l.disconnectVMFromManagementNetwork()
    print("Taking snapshots...")

    l.takeSnapshotordi(guacaComp)
    for name in listNewComp:
        l.takeSnapshotordi(name)
    print("Saving lab...")
    l.save()
    print(l)  
    shutil.rmtree("../Labs/" + l.name + '_add/') 
    os.system("rm ../lock")

if __name__ == "__main__":
    main()