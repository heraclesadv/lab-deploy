import sys
import shutil
import time
from Resources.labClass import *
from Resources.samples import *
from dotenv import load_dotenv
from openpyxl import load_workbook
from datetime import datetime
from Resources.Users import *

load_dotenv()

#Permet la création d'un LAB complet

def main():

    lockfile()
    
    if len(sys.argv) == 2 and (sys.argv[1] == "-h" or sys.argv[1] == "help" ):
            fichier = open("Resources/help/helpCreateLab.txt", 'r')
            print(fichier.read())
            fichier.close()
            sys.exit()
        

    #le deuxième argument doit faire référence au nom du Lab mais
    #si il n'est pas renseigné celui-ci sera demandé
    if len(sys.argv) <= 2 :
        name = input("Lab name: ").replace(" ", "")
    else :
        name = sys.argv[1] 
    
    
    l = lab(name)
    l.addComputer("ubuntu guacamole")
    if len(sys.argv) == 3 :
        wb = load_workbook("Resources/"+sys.argv[2])
        ws = wb.active
        ligne = 1
        test2= True
        while test2:
            c = ""
            colone = 1
            if ws.cell(row=ligne, column= colone).value != None and ws.cell(row=ligne, column= colone).value != "":
                test= True
                while test:
                    if ws.cell(row=ligne, column= colone).value != None and ws.cell(row=ligne, column= colone).value != "" :
                        c += ws.cell(row=ligne, column= colone).value
                    colone +=1
                    if ws.cell(row=ligne, column= colone).value != None and ws.cell(row=ligne, column= colone).value != "" :
                        c += " "
                    else :
                        test= False
                ca = c.split(" ")
                for d in ca:
                    if not d in ROLES:
                        print("Role "+d+" does not exist, computer not added.")
                        break
                else:
                    l.addComputer(c)
                    print("Computer added.")   
                ligne +=1
            else :
                test2 =False
    else :
        #Demande les VMs (ubuntu, win, ... ) à ajouter ainsi que leurs roles et verifie si ils existent
        while True:
            c = input("add >> ")
            if c.replace(" ", "") == "":
                break
            ca = c.split(" ")
            for d in ca:
                if not d in ROLES:
                    print("Role "+d+" does not exist, computer not added.")
                    break
            else:
                l.addComputer(c)
                print("Computer added.")

    l.createTfFiles()

    if not ask("Files created, go ? It's the right moment to edit the terraform files !"):
        print("Aborting...")
        l.network.freeHostOnlyNetwork()
        shutil.rmtree("../Labs/" + name)
        os.system("rm ../lock")
        return None

    print("Running Terraform ...")
    l.runTerraform()
    time.sleep(15)

    UsersCreation("Template_Users.xlsx",l.name)

    print("Creating ansible files...")
    l.createAnsibleFiles()
    print("Running ansible...")
    l.runAnsible()
    time.sleep(5)
    print("Done ! Cleaning...")
    l.cleanAnsibleFiles()
    print("Disconnecting management network...")
    l.disconnectVMFromManagementNetwork()
    time.sleep(15)
    print("Taking snapshots...")
    l.takeSnapshot()
    print("Saving lab...")
    l.save()
    print("Creation over, lab created.")
    print(l)
    os.system("rm ../lock")

if __name__ == "__main__":
    main()