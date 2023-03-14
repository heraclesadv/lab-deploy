import sys
from Resources.labClass import *
from Resources.samples import *
from dotenv import load_dotenv
from openpyxl import load_workbook
from os import path
import secrets
import string
import hashlib
import json

load_dotenv()

def genPW():
    alphabet = string.ascii_letters + string.digits + string.punctuation
    pwd_length = 12
    pwd = ''
    for i in range(pwd_length):
        pwd += ''.join(secrets.choice(alphabet))
    return pwd

#Permet la création de compte 

def deleteAllUsersLab(labName) :
    filename = "Resources/users.json"
    listObj = []
    with open(filename) as fp:
        json_object = json.load(fp)    
    for dict in json_object:
        if dict['Lab'] != labName:
            listObj.append(dict)
    
    with open(filename, 'w') as json_file:
        json.dump(listObj, json_file, 
                        indent=4,  
                        separators=(',',': '))

def UserModifPwd(id, labname) :
    
    filename = "Resources/users.json"
    f= open("Resources/Pwd/pwd"+ id +".txt","w+")
    listObj = []
    with open(filename) as fp:
        json_object = json.load(fp)    
    for dict in json_object:
        if dict['ID'] == id and dict['Lab'] == labname :
            pwd = genPW()
            result = hashlib.md5(pwd.encode())
            listObj.append({
            "Lab": dict["labname"],
            "ID": id,
            "PWD": result.hexdigest()
            })
            f.write("ID = " +dict["ID"] + " / Password = " + pwd  + " / MD5 = " + result.hexdigest()+"\n")
        else :
            listObj.append(dict)
    
    with open(filename, 'w') as json_file:
        json.dump(listObj, json_file, 
                        indent=4,  
                        separators=(',',': '))
    f.close()

def UsersCreation(file, labName) :
    
    wb = load_workbook("Resources/"+file)
    ws = wb.active
    
    filename = "Resources/users.json"
    listObj = []
    
    f= open("Resources/Pwd/pwdlist"+labName+".txt","w+")
    
    if path.isfile(filename) is False:
        raise Exception("File not found")
    with open(filename) as fp:
        listObj = json.load(fp)

    ligne = 1
    while ws.cell(row=ligne, column= 1).value != None and ws.cell(row=ligne, column= 1).value != "" :
        if ws.cell(row=ligne, column= 2).value == None or ws.cell(row=ligne, column= 2).value == "" : 
            ws.cell(row=ligne, column= 2).value =  genPW()
        # initializing string
        result = hashlib.md5(ws.cell(row=ligne, column= 2).value.encode())
        # printing the equivalent hexadecimal value.
        f.write("ID = " +ws.cell(row=ligne, column= 1).value + " / Password = " + ws.cell(row=ligne, column= 2).value  + " / MD5 = " +result.hexdigest()+"\n")
        listObj.append({
            "Lab": labName,
            "ID": ws.cell(row=ligne, column= 1).value,
            "PWD": result.hexdigest()
        })
        ligne += 1
    
    with open(filename, 'w') as json_file:
        json.dump(listObj, json_file, 
                        indent=4,  
                        separators=(',',': '))
    f.close()